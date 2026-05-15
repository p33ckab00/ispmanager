import csv
import hashlib
import io
import json
import zipfile
from datetime import date, datetime
from decimal import Decimal

from django.core.files.base import ContentFile
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone


def _display_value(value):
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if value is None:
        return ''
    return str(value)


def _workbook_value(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    return value


def canonical_csv_bytes(headers, rows):
    text = io.StringIO()
    writer = csv.writer(text)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([_display_value(value) for value in row])
    return text.getvalue().encode('utf-8')


def build_report_manifest(report_name, filename_base, headers, rows, filters, entity='', generated_by=''):
    csv_bytes = canonical_csv_bytes(headers, rows)
    return {
        'report_name': report_name,
        'entity': str(entity or ''),
        'generated_at': timezone.now().isoformat(),
        'generated_by': generated_by or '',
        'filters': filters,
        'row_count': len(rows),
        'columns': list(headers),
        'canonical_data': {
            'filename': f'{filename_base}.csv',
            'sha256': hashlib.sha256(csv_bytes).hexdigest(),
            'bytes': len(csv_bytes),
        },
        'notes': [
            'Hash is computed from canonical CSV data for this report run.',
            'A report archive record is stored when this export is generated.',
        ],
    }


def csv_report_response(filename, headers, rows, manifest):
    response = HttpResponse(canonical_csv_bytes(headers, rows), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['X-Report-Data-SHA256'] = manifest['canonical_data']['sha256']
    return response


def xlsx_report_response(filename, report_name, headers, rows, manifest):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    manifest_sheet = workbook.active
    manifest_sheet.title = 'Manifest'
    manifest_sheet.append(['Field', 'Value'])
    for key in ('report_name', 'entity', 'generated_at', 'generated_by', 'row_count'):
        manifest_sheet.append([key, manifest[key]])
    manifest_sheet.append(['canonical_csv_filename', manifest['canonical_data']['filename']])
    manifest_sheet.append(['canonical_csv_sha256', manifest['canonical_data']['sha256']])
    manifest_sheet.append(['filters', json.dumps(manifest['filters'], default=str, sort_keys=True)])

    report_sheet = workbook.create_sheet('Report')
    report_sheet.append(list(headers))
    for cell in report_sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='374151')
    for row in rows:
        report_sheet.append([_workbook_value(value) for value in row])
    report_sheet.freeze_panes = 'A2'

    for sheet in (manifest_sheet, report_sheet):
        for column_cells in sheet.columns:
            max_length = max(len(_display_value(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 48)

    buffer = io.BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['X-Report-Data-SHA256'] = manifest['canonical_data']['sha256']
    return response


def pdf_report_response(request, filename, report_name, headers, rows, filters, manifest, entity=''):
    html = render_to_string('accounting/report_export_pdf.html', {
        'report_name': report_name,
        'entity': entity,
        'headers': headers,
        'rows': rows,
        'filters': filters,
        'manifest': manifest,
    }, request=request)

    try:
        from xhtml2pdf import pisa
        pdf_buffer = io.BytesIO()
        status = pisa.CreatePDF(html.encode('utf-8'), dest=pdf_buffer, encoding='utf-8')
        if status.err:
            raise RuntimeError(f'PDF generation failed: {status.err}')
        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
    except Exception:
        response = HttpResponse(html, content_type='text/html')
        response['Content-Disposition'] = f'attachment; filename="{filename.replace(".pdf", ".html")}"'
    response['X-Report-Data-SHA256'] = manifest['canonical_data']['sha256']
    return response


def manifest_response(filename, manifest):
    payload = json.dumps(manifest, indent=2, default=str, sort_keys=True).encode('utf-8')
    response = HttpResponse(payload, content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['X-Report-Data-SHA256'] = manifest['canonical_data']['sha256']
    return response


def _record_report_archive(request, entity, export_format, response, filename, manifest, headers, rows):
    if not getattr(entity, 'pk', None):
        return None

    from apps.accounting.models import AccountingReportArchive

    safe_manifest = json.loads(json.dumps(manifest, default=str, sort_keys=True))
    payload = response.content
    file_sha256 = hashlib.sha256(payload).hexdigest()
    safe_manifest['generated_file'] = {
        'filename': filename,
        'content_type': response.get('Content-Type', ''),
        'sha256': file_sha256,
        'bytes': len(payload),
    }
    package_buffer = io.BytesIO()
    with zipfile.ZipFile(package_buffer, 'w', compression=zipfile.ZIP_DEFLATED) as package:
        package.writestr(filename, payload)
        if safe_manifest['canonical_data']['filename'] != filename:
            package.writestr(
                safe_manifest['canonical_data']['filename'],
                canonical_csv_bytes(headers, rows),
            )
        package.writestr(
            'manifest.json',
            json.dumps(safe_manifest, indent=2, default=str, sort_keys=True).encode('utf-8'),
        )
    package_payload = package_buffer.getvalue()
    archive = AccountingReportArchive(
        entity=entity,
        report_name=manifest['report_name'],
        export_format=export_format,
        filename=filename,
        content_type=response.get('Content-Type', ''),
        canonical_filename=manifest['canonical_data']['filename'],
        canonical_sha256=manifest['canonical_data']['sha256'],
        canonical_size=manifest['canonical_data']['bytes'],
        file_sha256=file_sha256,
        file_size=len(payload),
        row_count=manifest['row_count'],
        filters=safe_manifest.get('filters', {}),
        columns=safe_manifest.get('columns', []),
        manifest=safe_manifest,
        generated_by=request.user if getattr(request, 'user', None) and request.user.is_authenticated else None,
    )
    archive.archive_file.save(filename, ContentFile(payload), save=False)
    archive.package_file.save(f'{filename}.zip', ContentFile(package_payload), save=False)
    archive.package_sha256 = hashlib.sha256(package_payload).hexdigest()
    archive.package_size = len(package_payload)
    archive.save()
    response['X-Accounting-Report-Archive-ID'] = str(archive.pk)
    response['X-Accounting-Report-File-SHA256'] = archive.file_sha256
    return archive


def report_export_response(request, filename_base, report_name, headers, rows, filters, entity=''):
    export_format = (request.GET.get('format') or '').lower()
    if export_format not in ('csv', 'xlsx', 'pdf', 'manifest'):
        return None

    rows = list(rows)
    generated_by = request.user.get_username() if getattr(request, 'user', None) and request.user.is_authenticated else ''
    manifest = build_report_manifest(
        report_name,
        filename_base,
        headers,
        rows,
        filters,
        entity=entity,
        generated_by=generated_by,
    )
    if export_format == 'csv':
        filename = f'{filename_base}.csv'
        response = csv_report_response(filename, headers, rows, manifest)
    elif export_format == 'xlsx':
        filename = f'{filename_base}.xlsx'
        response = xlsx_report_response(filename, report_name, headers, rows, manifest)
    elif export_format == 'pdf':
        filename = f'{filename_base}.pdf'
        response = pdf_report_response(
            request,
            filename,
            report_name,
            headers,
            rows,
            filters,
            manifest,
            entity=entity,
        )
        if response['Content-Disposition'].endswith('.html"'):
            filename = filename.replace('.pdf', '.html')
    else:
        filename = f'{filename_base}-manifest.json'
        response = manifest_response(filename, manifest)
    _record_report_archive(request, entity, export_format, response, filename, manifest, headers, rows)
    return response
